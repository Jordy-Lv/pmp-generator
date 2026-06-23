#!/usr/bin/env python3
"""
PMP Generator — Célula 3  ·  v3.0
Herramienta de terminal para la rotación semanal de clientes PMP
(Mateo → Heiner → Estefania) del equipo de la Célula 3.

UI rica con  rich  +  questionary.  Lectura/escritura con  openpyxl.
Config persistente en  ~/.pmp_celula3.json

Lanzar:   pmp           (comando instalado en ~/.local/bin)
   o:     ~/.venvs/pmp/bin/python pmp_generator.py
"""
from __future__ import annotations

# ══════════════════════════════════════════════════════════════════════════════
#  0 · BOOTSTRAP DE DEPENDENCIAS
#     Si falta rich / questionary / openpyxl, se muestra un aviso, se instalan
#     automáticamente en el intérprete actual y el script se relanza una vez.
# ══════════════════════════════════════════════════════════════════════════════
import copy, importlib.util, json, os, subprocess, sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

_REQUERIDAS = ("rich", "questionary", "openpyxl")


def _faltantes() -> List[str]:
    return [m for m in _REQUERIDAS if importlib.util.find_spec(m) is None]


def _bootstrap_dependencias() -> None:
    # En el ejecutable empaquetado (PyInstaller) las dependencias ya van dentro:
    # no hay pip que invocar ni intérprete que relanzar, así que se omite.
    if getattr(sys, "frozen", False):
        return
    falt = _faltantes()
    if not falt:
        return
    print("\n  ⚠  Faltan librerías necesarias: " + ", ".join(falt))
    print("  ⏳  Instalándolas automáticamente con pip…\n")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", *falt])
    except Exception:
        print("\n  ❌  No se pudieron instalar automáticamente.")
        print("      Instálalas a mano y vuelve a ejecutar:\n")
        print(f"        {sys.executable} -m pip install {' '.join(falt)}\n")
        sys.exit(1)
    # Relanzar una sola vez para que los import del módulo encuentren todo.
    if os.environ.get("PMP_REEXEC") != "1":
        os.environ["PMP_REEXEC"] = "1"
        os.execv(sys.executable, [sys.executable, *sys.argv])
    # Ya se reintentó: si aún faltan, pip "instaló" pero este intérprete no las
    # ve (otro site-packages / PEP 668). Abortar con mensaje claro, no ImportError.
    if _faltantes():
        print("\n  ❌  Las librerías se instalaron pero no son visibles para este Python.")
        print("      Instálalas a mano en este intérprete y vuelve a ejecutar:\n")
        print(f"        {sys.executable} -m pip install {' '.join(falt)}\n")
        sys.exit(1)


_bootstrap_dependencias()

# A partir de aquí las dependencias están garantizadas.
import questionary
from questionary import Style as QStyle
from rich import box
from rich.console import Console, Group
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

console = Console()


# Estilo de los prompts de questionary (acento cian, coherente con el banner).
QS = QStyle([
    ("qmark",       "fg:#36c2ce bold"),
    ("question",    "bold"),
    ("pointer",     "fg:#36c2ce bold"),
    ("highlighted", "fg:#36c2ce bold"),
    ("selected",    "fg:#22a565"),
    ("answer",      "fg:#36c2ce bold"),
    ("instruction", "fg:#888888"),
])


class Volver(Exception):
    """El usuario canceló (Ctrl+C) — volver al menú principal."""


# ══════════════════════════════════════════════════════════════════════════════
#  1 · CONFIG PERSISTENTE  (~/.pmp_celula3.json)   ·  formato sin cambios
# ══════════════════════════════════════════════════════════════════════════════
CONFIG_PATH = Path.home() / ".pmp_celula3.json"
DEFAULT_CFG: dict = {
    "ruta_pmp":        "",
    "ruta_matriz":     "",
    # Orden del giro semanal: cada cliente de Célula 3 pasa de su encargado actual
    # al SIGUIENTE de esta lista (Heiner→Mateo→Estefania→Heiner).
    "rotacion_pmp":    ["Heiner Diaz", "Mateo Florez", "Estefania Sanabria"],
    "rotacion_n2":     ["Estefania Sanabria", "Heiner Diaz", "Mateo Florez"],
    "fecha_base_n2":   "2026-06-12",
    "rotacion_n3":     ["Carlos Barrera", "Santiago Amaya", "Adriano Carreño"],
    # Clientes que la Célula 3 rota entre sus 3 encargados. Se usan para LEER el
    # reparto de una semana por nombre de cliente (no por posición de columna), de
    # modo que la lectura aguante cambios de estructura del Excel.
    "clientes_celula3": [
        "La Riviera (APP)", "La Riviera (DB)", "HOMI Oracle", "HOMI SQLServer / MySQL",
        "Novaventa (DB2)", "Novaventa (SQLServer)", "EMI", "Acción Fiduciaria (Oracle)",
        "Acción Fiduciaria (SQLServer)", "Bancoldex (Oracle)",
    ],
    "horario_tarde":   ["La Riviera (APP)", "La Riviera (DB)", "HOMI Oracle", "HOMI SQLServer / MySQL"],
    "clientes_largos": ["HOMI Oracle", "HOMI SQLServer / MySQL"],
}


def cargar_cfg() -> dict:
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                return {**DEFAULT_CFG, **json.load(f)}
        except Exception:
            pass
    return DEFAULT_CFG.copy()


def guardar_cfg(cfg: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


FORMATOS_FECHA = ("%d/%m/%Y", "%Y-%m-%d", "%d%m%Y", "%d-%m-%Y")


def parse_fecha(raw: str) -> date:
    for fmt in FORMATOS_FECHA:
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError("Usa una fecha tipo DD/MM/AAAA")


# ══════════════════════════════════════════════════════════════════════════════
#  2 · LÓGICA DE NEGOCIO   ·  sin cambios funcionales
# ══════════════════════════════════════════════════════════════════════════════
def siguiente_lunes() -> date:
    hoy = date.today()
    dias = (7 - hoy.weekday()) % 7 or 7          # días hasta el próximo lunes
    return hoy + timedelta(days=dias)


def leer_dispo(lunes: date, ruta: str) -> Tuple[Optional[str], Optional[str]]:
    """Pestaña 'ROTACION DISPO CELULA 3 ' (con espacio final) de la Matriz."""
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    ws = wb["ROTACION DISPO CELULA 3 "]
    for row in ws.iter_rows(values_only=True):
        if len(row) < 5:
            continue                       # fila sin las 5 columnas esperadas
        ini, fin, n2, n3 = row[1], row[2], row[3], row[4]
        if hasattr(ini, "date") and hasattr(fin, "date"):
            if ini.date() <= lunes <= fin.date():
                # Preservar None en celdas vacías (str(None) daría el literal "None",
                # que es truthy y rompería el fallback / mostraría "None" en pantalla).
                return (str(n2) if n2 is not None else None,
                        str(n3) if n3 is not None else None)
    return None, None


def calcular_dispo_fallback(lunes: date, cfg: dict) -> str:
    """Fallback por cálculo cuando no hay Matriz: rota desde una fecha base.
    Tolera un config corrupto (lista vacía o fecha mal escrita) cayendo a los
    valores por defecto en vez de reventar con un traceback."""
    orden = cfg.get("rotacion_n2") or DEFAULT_CFG["rotacion_n2"]
    try:
        base = date.fromisoformat(str(cfg.get("fecha_base_n2", "")))
    except (ValueError, TypeError):
        base = date.fromisoformat(DEFAULT_CFG["fecha_base_n2"])
    idx = ((lunes - base).days // 7) % len(orden)
    return orden[idx]


def siguiente_disponible(nombre: str, rotacion: List[str], ausentes: List[str],
                         cobertura: Optional[Dict[str, str]] = None) -> str:
    """Siguiente consultor en la rotación. Si el siguiente está AUSENTE:
      · si `cobertura` dice quién lo reemplaza (elegido por el usuario), devuelve ese
        cubridor;
      · si no, salta al siguiente disponible (comportamiento por defecto).

    Es la ÚNICA regla de rotación del proyecto: la usan tanto `rotar()` (que mueve
    las carteras en memoria para el cuadro resumen) como `actualizar_control_pmp()`
    (que renombra los consultores dentro del Control real). Ambos reciben la MISMA
    `cobertura`, así los dos artefactos rotan idénticos y nunca divergen.
    """
    if nombre not in rotacion:
        return nombre
    n = len(rotacion)
    sig = (rotacion.index(nombre) + 1) % n
    if rotacion[sig] in ausentes and cobertura and rotacion[sig] in cobertura:
        return cobertura[rotacion[sig]]
    intentos = 0
    while rotacion[sig] in ausentes and intentos < n:
        sig = (sig + 1) % n
        intentos += 1
    return rotacion[sig]


def rotar(actuales: Dict[str, List[str]], ausentes: List[str], rotacion: List[str],
          cobertura: Optional[Dict[str, str]] = None) -> Dict[str, List[str]]:
    """Mateo→Heiner→Estefania→Mateo. Si el destino está ausente, lo cubre quien diga
    `cobertura` (elegido por el usuario) o, en su defecto, el siguiente disponible."""
    nuevas: Dict[str, List[str]] = {ing: [] for ing in rotacion}
    for ing in rotacion:
        clientes = actuales.get(ing, [])
        if clientes:
            nuevas[siguiente_disponible(ing, rotacion, ausentes, cobertura)].extend(clientes)
    return nuevas


# Clasificación de horario — extraída como helper para que el PREVIEW y el Excel
# muestren exactamente lo mismo (mismo cálculo, sin divergencias).
def horario_de_cliente(cliente: str, horario_tarde: set) -> str:
    return "Tarde (12:00–17:00)" if cliente in horario_tarde else "Mañana (7:00–12:00)"


def horario_predominante(clientes: List[str], cfg: dict) -> str:
    tarde = set(cfg["horario_tarde"])
    horarios = [horario_de_cliente(c, tarde) for c in clientes]
    return max(set(horarios), key=horarios.count) if horarios else "Mañana (7:00–12:00)"


def _guardar_wb(wb, destino) -> None:
    """Guarda el workbook de forma ATÓMICA y SIN dejar copias: escribe a un temporal
    en la misma carpeta y lo renombra sobre el destino (os.replace). Si el guardado
    falla a mitad, el archivo de destino anterior queda intacto (no se corrompe).
    Esto permite sobrescribir el mismo archivo de entrada con seguridad."""
    destino = str(destino)
    tmp = f"{destino}.tmp"
    try:
        wb.save(tmp)
        os.replace(tmp, destino)
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass
        raise


def generar_excel(lunes: date, asig: Dict[str, List[str]], n2: str, n3: Optional[str],
                  ausentes: List[str], cfg: dict, dir_destino: Path) -> str:
    """Genera el Excel formateado. Lo guarda en `dir_destino` (junto al PMP fuente)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rotacion      = cfg["rotacion_pmp"]
    largos        = set(cfg["clientes_largos"])
    dias          = [lunes + timedelta(i) for i in range(5)]
    dias_nom      = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    wb = Workbook(); ws = wb.active; ws.title = "Asignación PMP"

    def fill(h):       return PatternFill("solid", fgColor=h)
    def fnt(color="000000", bold=False, size=10): return Font(color=color, bold=bold, size=size)
    def aln(h="center", v="center"): return Alignment(horizontal=h, vertical=v, wrap_text=True)
    t  = Side(border_style="thin", color="BBBBBB")
    br = Border(left=t, right=t, top=t, bottom=t)
    N  = 2 + len(dias)

    def hdr(cell, txt, bg, fg="FFFFFF", sz=10, bold=True):
        cell.value = txt; cell.fill = fill(bg)
        cell.font  = fnt(fg, bold, sz); cell.alignment = aln(); cell.border = br

    ws.merge_cells(f"A1:{get_column_letter(N)}1")
    hdr(ws["A1"], f"📋  PMP SEMANAL — {lunes.strftime('%d/%m/%Y')} al {(lunes+timedelta(4)).strftime('%d/%m/%Y')}",
        "C00000", sz=12)
    ws.row_dimensions[1].height = 30

    hdr(ws["A2"], "Consultor", "C00000"); hdr(ws["B2"], "Horario", "C00000")
    for i, (d_, nom) in enumerate(zip(dias, dias_nom)):
        hdr(ws[f"{get_column_letter(3+i)}2"], f"{nom}\n{d_.strftime('%d/%m/%Y')}", "E74C3C")
    ws.row_dimensions[2].height = 38

    fila = 3
    colores_fila = ["FDEDEC", "FFFFFF"]
    for idx, ing in enumerate(rotacion):
        ausente  = ing in ausentes
        clientes = asig.get(ing, [])
        cf       = "D5D8DC" if ausente else colores_fila[idx % 2]
        ws.row_dimensions[fila].height = max(32, 18 * max(len(clientes), 1))

        horario_display = horario_predominante(clientes, cfg)

        ca = ws[f"A{fila}"]
        ca.value = ing; ca.fill = fill(cf)
        ca.font  = fnt(bold=True, color="CC0000" if ausente else "000000")
        ca.alignment = aln("left"); ca.border = br

        cb = ws[f"B{fila}"]
        cb.value = horario_display; cb.fill = fill(cf)
        cb.font  = fnt(); cb.alignment = aln(); cb.border = br

        for j in range(5):
            col = get_column_letter(3 + j); cc = ws[f"{col}{fila}"]
            cc.border = br; cc.alignment = aln("left")
            if ausente:
                cc.value = "\n".join(f"⚠️ {c_} — REASIGNAR" for c_ in clientes) or "—"
                cc.fill = fill("FADBD8"); cc.font = fnt("922B21", True)
            else:
                lineas = [c_ + (" ⚠️ PMP largo" if c_ in largos else "") for c_ in clientes]
                cc.value = "\n".join(lineas) or "—"
                cc.fill  = fill("FEF9E7" if any(c_ in largos for c_ in clientes) else cf)
                cc.font  = fnt()
        fila += 1

    for texto, color in [
        (f"🌙  DISPO NOCTURNA N2 (V-V 10pm–7am): {n2}", "FFC300"),
        *([(f"🆘  ESCALAMIENTO N3 (cambios / desbordes): {n3}", "F0B27A")] if n3 else []),
        *([(f"⚠️  Ausentes: {', '.join(ausentes)} — revisar reasignación", "E74C3C")] if ausentes else []),
    ]:
        ws.merge_cells(f"A{fila}:{get_column_letter(N)}{fila}")
        cc = ws[f"A{fila}"]
        cc.value = texto; cc.fill = fill(color)
        cc.font  = fnt("FFFFFF" if color == "E74C3C" else "000000", True)
        cc.alignment = aln(); cc.border = br
        ws.row_dimensions[fila].height = 22; fila += 1

    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 22
    for i in range(5):
        ws.column_dimensions[get_column_letter(3+i)].width = 32

    # Hoja Config — parámetros de esta generación
    ws2 = wb.create_sheet("Config")
    ws2["A1"] = "Parámetros de esta generación"; ws2["A1"].font = fnt(bold=True)
    fin = (lunes + timedelta(4)).strftime("%d/%m/%Y")
    for r_, (k, v) in enumerate([
        ("Semana", f"{lunes.strftime('%d/%m/%Y')} → {fin}"),
        ("Generado el", date.today().strftime("%d/%m/%Y")),
        ("Dispo N2", n2), ("Dispo N3", n3 or "N/A"),
        ("Ausentes", ", ".join(ausentes) or "Ninguno"),
    ], 2):
        ws2[f"A{r_}"] = k; ws2[f"B{r_}"] = v; ws2[f"A{r_}"].font = fnt(bold=True)
    ws2.column_dimensions["A"].width = 16; ws2.column_dimensions["B"].width = 50

    # Nombre FIJO (sin fecha): se sobrescribe cada semana, no acumula copias.
    destino = Path(dir_destino) / "PMP_Semana.xlsx"
    _guardar_wb(wb, destino)
    return str(destino)


# ══════════════════════════════════════════════════════════════════════════════
#  3 · UI  ·  banner, helpers visuales y prompts de questionary
# ══════════════════════════════════════════════════════════════════════════════
def limpiar() -> None:
    console.clear()


def banner() -> None:
    limpiar()
    titulo = Text("📋  PMP GENERATOR", style="bold white")
    titulo.append("  ·  Célula 3", style="bold cyan")
    sub = Text("Rotación semanal  ·  Mateo → Heiner → Estefania", style="dim")
    console.print(Panel(Group(titulo, sub), box=box.DOUBLE, border_style="cyan",
                        padding=(0, 2), expand=False))


def _ask(pregunta):
    """Ejecuta un prompt de questionary; si el usuario cancela (Ctrl+C) → Volver.
    Usa unsafe_ask() porque ask() se traga el Ctrl+C e imprime 'Cancelled by user'."""
    try:
        resp = pregunta.unsafe_ask()
    except (KeyboardInterrupt, EOFError):
        raise Volver()
    if resp is None:
        raise Volver()
    return resp


def confirmar(mensaje: str, default: bool = True) -> bool:
    return _ask(questionary.confirm(mensaje, default=default, style=QS, auto_enter=False))


def pedir_fecha(prompt: str, default: Optional[date] = None) -> date:
    """Pide una fecha con varios formatos aceptados; reintenta hasta que sea válida."""
    default_str = default.strftime("%d/%m/%Y") if default else ""
    while True:
        raw = _ask(questionary.text(
            prompt, default=default_str, style=QS,
            instruction="(DD/MM/AAAA · Enter para el valor sugerido)")).strip()
        try:
            return parse_fecha(raw)
        except ValueError:
            pass
        console.print(f"  [red]✗[/] Formato no válido: '{raw}'. Usa DD/MM/AAAA.")


def seleccionar_ausentes(ingenieros: List[str]) -> List[str]:
    """Checkbox con flechas + espacio. Sin marcar = nadie ausente."""
    return _ask(questionary.checkbox(
        "¿Alguien ausente esta semana?  (Espacio para marcar · Enter para confirmar)",
        choices=ingenieros, style=QS))


def pedir_cobertura(rotacion: List[str], ausentes: List[str]) -> Dict[str, str]:
    """Por cada ausente, pregunta QUIÉN cubre los PMP que esta semana le tocarían y
    devuelve {ausente: cubridor}. Así la reasignación la decide el usuario (eligiendo
    entre los presentes) en vez de saltar automáticamente al siguiente de la rotación.
    Si solo queda un presente, lo asigna sin preguntar."""
    presentes = [c for c in rotacion if c not in ausentes]
    cobertura: Dict[str, str] = {}
    if not presentes:
        return cobertura
    for aus in [a for a in ausentes if a in rotacion]:
        if len(presentes) == 1:
            cobertura[aus] = presentes[0]
            console.print(f"  [dim]· {presentes[0]} cubre a {aus} (único disponible).[/]")
        else:
            cobertura[aus] = _ask(questionary.select(
                f"¿Quién cubre los PMP de {aus} esta semana?", choices=presentes, style=QS))
    return cobertura


# ── Auto-detección de los Excel ───────────────────────────────────────────────
DIRS_BUSQUEDA = [Path.home() / "Downloads", Path.home() / "Desktop", Path.home() / "Documents"]


def buscar_excels(patron: str) -> List[Path]:
    """Busca *<patron>*.xlsx (SIN distinguir mayúsculas) en Downloads/Desktop/
    Documents y un nivel de subcarpetas."""
    encontrados: dict[str, Path] = {}
    aguja = patron.lower()

    def coincide(nombre: str) -> bool:
        n = nombre.lower()
        return aguja in n and n.endswith(".xlsx") and not nombre.startswith("~$")

    for base in DIRS_BUSQUEDA:
        if not base.is_dir():
            continue
        # nivel 0 (la carpeta base) + nivel 1 (sus subcarpetas directas)
        candidatos = list(base.iterdir())
        for sub in [d for d in list(candidatos) if d.is_dir()]:
            try:
                candidatos.extend(sub.iterdir())
            except OSError:
                continue          # subcarpeta sin permisos: se ignora
        for p in candidatos:
            if p.is_file() and coincide(p.name):
                encontrados[str(p.resolve())] = p
    return sorted(encontrados.values(), key=lambda p: (len(p.parts), p.name.lower()))


def _etiqueta_archivo(p: Path) -> str:
    try:
        rel = p.relative_to(Path.home())
        return f"~/{rel}"
    except ValueError:
        return str(p)


def resolver_ruta(cfg: dict, clave: str, patron: str, descripcion: str,
                  requerido: bool = True) -> Optional[str]:
    """
    Devuelve una ruta válida para `clave`:
      1. la guardada en config, si el archivo existe;
      2. auto-detección — 1 resultado se usa directo, varios se eligen con select;
      3. si no hay ninguno, se pide la ruta a mano (autocompletado de path).
    El resultado se persiste en la config.
    """
    actual = cfg.get(clave, "")
    if actual and Path(actual).is_file():
        return actual

    halladas = buscar_excels(patron)

    if len(halladas) == 1:
        ruta = str(halladas[0].resolve())
        console.print(f"  [green]✓[/] {descripcion} detectado: [cyan]{_etiqueta_archivo(halladas[0])}[/]")
        cfg[clave] = ruta; guardar_cfg(cfg)
        return ruta

    if len(halladas) > 1:
        opciones = [questionary.Choice(title=_etiqueta_archivo(p), value=str(p.resolve()))
                    for p in halladas]
        opciones.append(questionary.Choice(title="✎  Escribir la ruta manualmente…", value="__manual__"))
        elegida = _ask(questionary.select(
            f"Se encontraron varios candidatos para {descripcion}:", choices=opciones, style=QS))
        if elegida != "__manual__":
            cfg[clave] = elegida; guardar_cfg(cfg)
            return elegida

    # 0 resultados (o eligió escribir manualmente)
    if not requerido and len(halladas) == 0:
        if not confirmar(f"No se encontró {descripcion}. ¿Indicar la ruta manualmente?", default=False):
            return None
    return _pedir_ruta_manual(cfg, clave, descripcion, requerido)


def _pedir_ruta_manual(cfg: dict, clave: str, descripcion: str, requerido: bool) -> Optional[str]:
    while True:
        raw = _ask(questionary.path(f"Ruta a {descripcion}:", style=QS)).strip()
        if not raw and not requerido:
            return None
        ruta = str(Path(raw).expanduser())
        if Path(ruta).is_file() and ruta.lower().endswith(".xlsx"):
            cfg[clave] = ruta; guardar_cfg(cfg)
            return ruta
        console.print(f"  [red]✗[/] No es un .xlsx válido: '{raw}'")
        if not requerido and not confirmar("¿Intentar de nuevo?", default=True):
            return None


def auto_detectar_inicio(cfg: dict) -> None:
    """Al arrancar: rellena en silencio las rutas únicas y muestra un panel-resumen."""
    lineas: List[Text] = []
    # No se abre ningún Excel aquí (el menú debe aparecer al instante): solo se
    # resuelven rutas. La preferencia por el Control con la semana más avanzada se
    # aplica al GENERAR (modo_pmp), que es donde importa.
    for clave, patron, desc in [("ruta_pmp", "PMP", "Control de Gestión PMP"),
                                 ("ruta_matriz", "Matriz", "Matriz Unificada")]:
        actual = cfg.get(clave, "")
        if actual and Path(actual).is_file():
            lineas.append(Text.assemble(("✓ ", "green"), (f"{desc}: ", "bold"),
                                        (_etiqueta_archivo(Path(actual)), "cyan")))
            continue
        halladas = buscar_excels(patron)
        if len(halladas) == 1:
            cfg[clave] = str(halladas[0].resolve()); guardar_cfg(cfg)
            lineas.append(Text.assemble(("✓ ", "green"), (f"{desc}: ", "bold"),
                                        (_etiqueta_archivo(halladas[0]), "cyan")))
        elif len(halladas) > 1:
            lineas.append(Text.assemble(("• ", "yellow"), (f"{desc}: ", "bold"),
                                        (f"{len(halladas)} candidatos (se elegirá al generar)", "dim")))
        else:
            lineas.append(Text.assemble(("• ", "yellow"), (f"{desc}: ", "bold"),
                                        ("sin detectar (se pedirá al usar)", "dim")))
    console.print(Panel(Group(*lineas), title="🔍  Detección de archivos",
                        border_style="cyan", box=box.ROUNDED, expand=False, padding=(0, 2)))


# ── Preview de la rotación ─────────────────────────────────────────────────────
def tabla_preview(lunes: date, nuevas: Dict[str, List[str]], ausentes: List[str],
                  n2: str, n3: Optional[str], fuente_dispo: str, cfg: dict) -> Panel:
    largos = set(cfg["clientes_largos"])
    fin = (lunes + timedelta(4)).strftime("%d/%m/%Y")

    tabla = Table(box=box.SIMPLE_HEAVY, expand=False, show_lines=True,
                  header_style="bold white on red3")
    tabla.add_column("Ingeniero", style="bold", no_wrap=True)
    tabla.add_column("Horario", justify="center")
    tabla.add_column("Clientes asignados")

    for ing in cfg["rotacion_pmp"]:
        clientes = nuevas.get(ing, [])
        if ing in ausentes:
            ing_cell = Text(f"{ing}  (AUSENTE)", style="bold red")
            cli_txt  = Text("\n".join(f"⚠ {c} — REASIGNAR" for c in clientes) or "—", style="red")
            horario  = Text("—", style="dim")
        else:
            ing_cell = Text(ing, style="bold")
            horario  = Text(horario_predominante(clientes, cfg).replace(" (", "\n("))
            partes = []
            for c in clientes:
                t = Text(c)
                if c in largos:
                    t.append("  ⚠ PMP largo", style="yellow")
                partes.append(t)
            cli_txt = Text("\n").join(partes) if partes else Text("(sin clientes)", style="dim")
        tabla.add_row(ing_cell, horario, cli_txt)

    pie = Text()
    pie.append("🌙  Dispo N2 (V-V 10pm–7am): ", style="bold")
    pie.append(f"{n2}", style="yellow")
    pie.append(f"   [{fuente_dispo}]\n", style="dim")
    if n3:
        pie.append("🆘  Escalamiento N3: ", style="bold"); pie.append(f"{n3}\n", style="orange3")
    if ausentes:
        pie.append("⚠  Ausentes: ", style="bold red")
        pie.append(", ".join(ausentes), style="red")

    cuerpo = Group(tabla, Text(), pie)
    return Panel(cuerpo, title=f"👁  Vista previa — semana {lunes.strftime('%d/%m/%Y')} → {fin}",
                 border_style="green", box=box.ROUNDED, padding=(1, 2), expand=False)


def abrir_archivo(ruta: str) -> None:
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", ruta], check=False)
        elif os.name == "nt":
            os.startfile(ruta)            # type: ignore[attr-defined]
        else:
            subprocess.run(["xdg-open", ruta], check=False)
    except Exception as e:
        console.print(f"  [yellow]⚠[/] No se pudo abrir automáticamente: {e}")


def _siguiente_en_rotacion(nombre: str, rotacion: List[str]) -> str:
    return rotacion[(rotacion.index(nombre) + 1) % len(rotacion)] if nombre in rotacion else nombre


def _fecha_excel(d: date) -> datetime:
    return datetime.combine(d, datetime.min.time())


# ── Festivos de Colombia (deterministas, sin internet) ───────────────────────────
def _domingo_pascua(anio: int) -> date:
    """Domingo de Pascua por el algoritmo de Butcher (calendario gregoriano)."""
    a = anio % 19
    b, c = divmod(anio, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    el = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * el) // 451
    mes, dia = divmod(h + el - 7 * m + 114, 31)
    return date(anio, mes, dia + 1)


def festivos_colombia(anio: int) -> set:
    """Conjunto de fechas festivas de Colombia en `anio` (ley Emiliani + Pascua)."""
    def lunes_siguiente(d: date) -> date:
        return d + timedelta(days=(7 - d.weekday()) % 7)   # 0 si ya es lunes

    fest = {date(anio, m, d) for m, d in [(1, 1), (5, 1), (7, 20), (8, 7), (12, 8), (12, 25)]}
    # Trasladables al lunes siguiente
    for m, d in [(1, 6), (3, 19), (6, 29), (8, 15), (10, 12), (11, 1), (11, 11)]:
        fest.add(lunes_siguiente(date(anio, m, d)))
    # Móviles ligados a la Pascua
    p = _domingo_pascua(anio)
    fest.add(p - timedelta(days=3))                       # Jueves Santo
    fest.add(p - timedelta(days=2))                       # Viernes Santo
    fest.add(lunes_siguiente(p + timedelta(days=39)))     # Ascensión
    fest.add(lunes_siguiente(p + timedelta(days=60)))     # Corpus Christi
    fest.add(lunes_siguiente(p + timedelta(days=68)))     # Sagrado Corazón
    return fest


def es_festivo(d: date) -> bool:
    return d in festivos_colombia(d.year)


def _cliente_de_fila(ws, r: int, cols_consultor: set) -> Optional[str]:
    """Cliente que atiende la fila `r`: el primer texto que no sea consultor,
    horario, fecha, '-' ni 'FESTIVO'. Sirve para rellenar la columna de un día que
    dejó de ser festivo (donde antes había 'FESTIVO')."""
    for c in range(1, ws.max_column + 1):
        if c in cols_consultor:
            continue
        v = ws.cell(r, c).value
        if not isinstance(v, str):
            continue
        s = v.strip()
        if not s or s in ("-", "FESTIVO", "Consultor", "HORA DE ENVIO"):
            continue
        if s.startswith("Mañana") or s.startswith("Tarde"):
            continue
        return s
    return None


def _es_celda_festivo(celda) -> bool:
    return isinstance(celda.value, str) and celda.value.strip().upper() == "FESTIVO"


def _merge_festivo_en_col(ws, col: int, fila_desde: int, fila_hasta: int):
    """Merge vertical 'FESTIVO' en la columna `col` dentro del bloque, o None."""
    for m in ws.merged_cells.ranges:
        if m.min_col == col == m.max_col and fila_desde <= m.min_row <= fila_hasta:
            if _es_celda_festivo(ws.cell(m.min_row, col)):
                return m
    return None


def _celda_festivo_ref(ws):
    """Una celda 'FESTIVO' existente (esquina de su merge), para clonar su formato
    al marcar días festivos nuevos. None si no hay ninguna en la hoja."""
    for m in ws.merged_cells.ranges:
        if _es_celda_festivo(ws.cell(m.min_row, m.min_col)):
            return ws.cell(m.min_row, m.min_col)
    return None


def _rango_datos_bloque(ws, target_row: int, alto: int, date_cols: List[int]) -> Tuple[int, int]:
    """Filas de datos del bloque (las que abarca un 'FESTIVO' de columna completa):
    si el bloque ya trae algún FESTIVO se usa su extensión; si no, las filas que
    tengan algún cliente en las columnas-día."""
    fests = [m for m in ws.merged_cells.ranges
             if m.min_col == m.max_col and target_row < m.min_row <= target_row + alto
             and _es_celda_festivo(ws.cell(m.min_row, m.min_col))]
    if fests:
        return min(m.min_row for m in fests), max(m.max_row for m in fests)
    filas = [r for r in range(target_row + 1, target_row + alto)
             if any(isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip()
                    for c in date_cols)]
    return target_row + 1, max(filas, default=target_row + alto - 1)


def _quitar_festivo_col(ws, col: int, merge, cols_cons: set) -> None:
    """Deshace un merge 'FESTIVO' en `col` y rellena cada fila con su cliente,
    clonando el formato de la columna del mismo cliente (como un día normal)."""
    filas = list(range(merge.min_row, merge.max_row + 1))
    ws.unmerge_cells(str(merge))
    for r in filas:
        cli = _cliente_de_fila(ws, r, cols_cons)
        d = ws.cell(r, col)
        d.value = cli
        # Al deshacer la combinación el día queda sin estilo: copiar el formato de la
        # columna del MISMO cliente (negrita, centrado, color) para que se vea igual.
        ref_col = next((c for c in range(col + 1, ws.max_column + 1)
                        if c not in cols_cons and ws.cell(r, c).value == cli), None)
        if ref_col:
            ref = ws.cell(r, ref_col)
            d._style = copy.copy(ref._style)
            d.font = copy.copy(ref.font)
            d.fill = copy.copy(ref.fill)
            d.border = copy.copy(ref.border)
            d.alignment = copy.copy(ref.alignment)
            d.number_format = ref.number_format


def _poner_festivo_col(ws, col: int, fila_ini: int, fila_fin: int, ref=None) -> None:
    """Marca la columna `col` como 'FESTIVO': deshace merges previos en el rango,
    limpia los valores, combina fila_ini..fila_fin y escribe 'FESTIVO', clonando el
    formato de `ref` (una celda FESTIVO existente) o, si no hay, negrita + centrado."""
    for m in list(ws.merged_cells.ranges):
        if m.min_col == col == m.max_col and not (m.max_row < fila_ini or m.min_row > fila_fin):
            ws.unmerge_cells(str(m))
    for r in range(fila_ini, fila_fin + 1):
        ws.cell(r, col).value = None
    if fila_fin > fila_ini:
        ws.merge_cells(start_row=fila_ini, start_column=col, end_row=fila_fin, end_column=col)
    celda = ws.cell(fila_ini, col)
    celda.value = "FESTIVO"
    if ref is not None:
        celda._style = copy.copy(ref._style)
        celda.font = copy.copy(ref.font)
        celda.fill = copy.copy(ref.fill)
        celda.border = copy.copy(ref.border)
        celda.alignment = copy.copy(ref.alignment)
        celda.number_format = ref.number_format
    else:
        from openpyxl.styles import Font, Alignment
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")


def _headers_semana_pmp(ws) -> List[Tuple[int, date]]:
    # Itera por filas (streaming) en vez de leer celda a celda con ws.cell(r, c):
    # en modo read_only el acceso ALEATORIO por celda relee el archivo en cada
    # llamada (O(n) por celda), lo que hacía que abrir el Control tardara ~10 s.
    # iter_rows es secuencial y baja eso a ~0.2 s. row[1]=col B, row[3]=col D.
    headers: List[Tuple[int, date]] = []
    for row in ws.iter_rows():
        if len(row) < 4:
            continue
        celda_fecha = row[3]            # columna D
        fecha = celda_fecha.value
        # OJO read_only: las celdas vacías son EmptyCell y NO tienen .row; por eso
        # se toma el nº de fila de la celda de la fecha (que sí tiene valor) y solo
        # cuando hay coincidencia.
        if hasattr(fecha, "date") and str(row[1].value).strip() == "Consultor":
            headers.append((celda_fecha.row, fecha.date()))
    return headers


def _espaciado_bloques(headers: List[Tuple[int, date]]) -> int:
    """Filas que ocupa cada bloque semanal = distancia entre los dos últimos
    encabezados. El archivo real arrancó con bloques de 27-29 filas y se estabilizó
    en 24; usar la última distancia hace que copiar/pegar la semana nueva use el
    tamaño real vigente en lugar de un número cableado."""
    filas = sorted(r for r, _ in headers)
    if len(filas) >= 2:
        return filas[-1] - filas[-2]
    return 24


def leer_reparto_c3(ws, lunes: date, clientes_c3: List[str],
                    consultores: List[str]) -> Tuple[Dict[str, List[str]], List[str]]:
    """Lee el reparto de los clientes de Célula 3 para la semana `lunes`, robusto a
    la posición de las columnas.

    En cada fila del bloque busca el cliente de C3 (por nombre) y lo asigna al
    consultor de C3 que aparezca MÁS A LA DERECHA en esa misma fila — esa es la
    tabla-resumen de Célula 3, que en este archivo se ha movido de la columna M a la
    L. Devuelve ({consultor: [clientes]}, [clientes_no_encontrados]).
    """
    reparto: Dict[str, List[str]] = {c: [] for c in consultores}
    set_cons = set(consultores)
    set_cli = set(clientes_c3)
    headers = _headers_semana_pmp(ws)
    fila_ini = next((r for r, f in headers if f == lunes), None)
    if fila_ini is None:
        return reparto, list(clientes_c3)

    espaciado = _espaciado_bloques(headers)
    siguientes = [r for r, _ in headers if r > fila_ini]
    fila_fin = min(siguientes) if siguientes else fila_ini + espaciado
    maxcol = ws.max_column

    vistos: set = set()
    for row in range(fila_ini + 1, fila_fin):
        # cliente de C3 en la fila (puede repetirse por día; basta uno)
        cliente = None
        for col in range(1, maxcol + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.strip() in set_cli:
                cliente = v.strip()
                break
        if not cliente:
            continue
        # consultor de C3 más a la derecha (= tabla resumen de Célula 3)
        consultor = None
        for col in range(maxcol, 0, -1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.strip() in set_cons:
                consultor = v.strip()
                break
        if consultor and cliente not in reparto[consultor]:
            reparto[consultor].append(cliente)
            vistos.add(cliente)

    faltantes = [c for c in clientes_c3 if c not in vistos]
    return reparto, faltantes


def leer_reparto_actual(ruta_pmp: str, lunes: date, cfg: dict) -> Tuple[Dict[str, List[str]], List[str]]:
    """Abre el Control y lee el reparto de Célula 3 de la semana `lunes`.
    Devuelve ({consultor: [clientes]}, [clientes no ubicados])."""
    from openpyxl import load_workbook
    try:
        ws = load_workbook(ruta_pmp, data_only=True)["2026"]
    except Exception:
        return {c: [] for c in cfg["rotacion_pmp"]}, list(cfg.get("clientes_celula3", []))
    return leer_reparto_c3(ws, lunes, cfg.get("clientes_celula3", []), cfg["rotacion_pmp"])


def siguiente_lunes_en_pmp(ruta_pmp: str) -> Optional[date]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta_pmp, read_only=True, data_only=False)
    ws = wb["2026"]
    headers = _headers_semana_pmp(ws)
    return max((lunes for _, lunes in headers), default=None) + timedelta(days=7) if headers else None


def _ultima_semana_pmp(ruta: str) -> Optional[date]:
    """Última semana (lunes) registrada en la pestaña '2026' de un Control, o None
    si no se puede leer o no tiene semanas. Tolerante: nunca lanza."""
    try:
        from openpyxl import load_workbook
        ws = load_workbook(ruta, read_only=True, data_only=False)["2026"]
        headers = _headers_semana_pmp(ws)
    except Exception:
        return None
    return max((lunes for _, lunes in headers), default=None)


def control_mas_reciente(cfg: dict) -> Optional[str]:
    """Entre todos los Control candidatos —los detectados en las carpetas habituales,
    los de la carpeta del Control configurado (ahí viven las salidas «_actualizado_»)
    y el propio configurado—, devuelve la ruta del que tiene la semana MÁS avanzada.
    Así el flujo siempre parte de la copia más actualizada aunque haya varias copias.
    None si ninguno es legible."""
    candidatos: Dict[str, Path] = {}
    for p in buscar_excels("PMP"):
        candidatos[str(p.resolve())] = p
    actual = cfg.get("ruta_pmp", "")
    if actual and Path(actual).is_file():
        ap = Path(actual)
        candidatos[str(ap.resolve())] = ap
        try:
            for q in ap.parent.glob("*.xlsx"):
                if "pmp" in q.name.lower() and not q.name.startswith("~$"):
                    candidatos[str(q.resolve())] = q
        except OSError:
            pass
    mejor, mejor_sem = None, None
    for ruta in candidatos:
        sem = _ultima_semana_pmp(ruta)
        if sem is None:
            continue
        if mejor_sem is None or sem > mejor_sem:
            mejor, mejor_sem = ruta, sem
    return mejor


def diagnostico_fecha(ruta_pmp: str, lunes_sig: date) -> Tuple[bool, Optional[str], Optional[date]]:
    """Comprueba —sin escribir nada— si la semana pedida encaja en el Control.
    Es determinista (no IA): la coherencia de fechas es una resta exacta.
    Devuelve (ok, mensaje, sugerida), donde `sugerida` es la siguiente semana que
    falta. ok=False si la semana ya existe o si no es la consecutiva esperada."""
    try:
        from openpyxl import load_workbook
        ws = load_workbook(ruta_pmp, read_only=True, data_only=False)["2026"]
        headers = _headers_semana_pmp(ws)
    except Exception:
        return True, None, None          # si no se puede leer, lo maneja el flujo normal
    if not headers:
        return True, None, None
    fechas = {f for _, f in headers}
    sugerida = max(fechas) + timedelta(days=7)
    if lunes_sig in fechas:
        return (False,
                f"  [red]✗[/] La semana {lunes_sig.strftime('%d/%m/%Y')} ya está en el Control "
                f"(esa tabla ya se creó). La siguiente que falta es "
                f"[cyan]{sugerida.strftime('%d/%m/%Y')}[/].",
                sugerida)
    if lunes_sig != sugerida:
        return (False,
                f"  [yellow]⚠[/] Las semanas deben ir consecutivas: la siguiente esperada es "
                f"[cyan]{sugerida.strftime('%d/%m/%Y')}[/], no {lunes_sig.strftime('%d/%m/%Y')}.",
                sugerida)
    return True, None, sugerida


def _copiar_bloque(ws, fila_origen: int, fila_destino: int, alto: int = 21,
                  col_min: int = 1, col_max: int = 15) -> None:
    offset = fila_destino - fila_origen

    for row in range(fila_origen, fila_origen + alto):
        ws.row_dimensions[row + offset].height = ws.row_dimensions[row].height
        for col in range(col_min, col_max + 1):
            origen = ws.cell(row, col)
            destino = ws.cell(row + offset, col)
            destino.value = origen.value
            if origen.has_style:
                destino._style = copy.copy(origen._style)
            destino.number_format = origen.number_format
            destino.font = copy.copy(origen.font)
            destino.fill = copy.copy(origen.fill)
            destino.border = copy.copy(origen.border)
            destino.alignment = copy.copy(origen.alignment)
            destino.protection = copy.copy(origen.protection)

    for merged in list(ws.merged_cells.ranges):
        if (fila_origen <= merged.min_row <= fila_origen + alto - 1 and
                fila_origen <= merged.max_row <= fila_origen + alto - 1):
            ws.merge_cells(
                start_row=merged.min_row + offset,
                start_column=merged.min_col,
                end_row=merged.max_row + offset,
                end_column=merged.max_col,
            )


def actualizar_control_pmp(ruta_pmp: str, lunes: date, ausentes: List[str], cfg: dict,
                           salida: Optional[str] = None,
                           cobertura: Optional[Dict[str, str]] = None) -> str:
    """Copia el último bloque semanal del Control PMP debajo y rota Célula 3.

    No reconstruye la tabla: conserva el formato real de Excel y solo toca el
    nuevo bloque copiado. La rotación de consultores usa `siguiente_disponible`,
    la MISMA regla que `rotar()`, de modo que el Control actualizado coincide con
    el cuadro resumen incluso cuando hay ausentes (el cliente del ausente queda a
    cargo del siguiente disponible; el aviso visual "REASIGNAR" vive solo en el
    cuadro resumen). Escribe en `salida` (si se omite, junto al original con sufijo
    `_actualizado_<fecha>`). En el flujo normal `modo_pmp` pasa `salida=ruta_pmp`,
    de modo que SOBRESCRIBE el mismo Control in-place; el guardado es atómico
    (`_guardar_wb`: escribe a un temporal y reemplaza), así un fallo a mitad no deja
    el archivo a medias.
    """
    from openpyxl import load_workbook

    wb = load_workbook(ruta_pmp)
    ws = wb["2026"]
    headers = _headers_semana_pmp(ws)
    if not headers:
        raise ValueError("No encontré bloques semanales en la hoja 2026.")

    if any(fecha == lunes for _, fecha in headers):
        raise ValueError(f"Ya existe un bloque para {lunes.strftime('%d/%m/%Y')}.")

    source_row, source_lunes = max((h for h in headers if h[1] < lunes), default=headers[-1])
    if lunes != source_lunes + timedelta(days=7):
        raise ValueError(
            f"La siguiente semana esperada después de {source_lunes.strftime('%d/%m/%Y')} "
            f"es {(source_lunes + timedelta(days=7)).strftime('%d/%m/%Y')}."
        )

    # Tamaño real del bloque (header→header) en vez de un alto cableado, para copiar
    # la semana COMPLETA sin perder filas finales (p. ej. las de Terpel).
    alto = _espaciado_bloques(headers)
    target_row = source_row + alto
    if any(ws.cell(r, c).value is not None
           for r in range(target_row, target_row + alto) for c in range(1, 16)):
        raise ValueError(f"La zona destino desde la fila {target_row} no está vacía.")

    _copiar_bloque(ws, source_row, target_row, alto=alto)

    date_cols = [
        c for c in range(1, ws.max_column + 1)
        if hasattr(ws.cell(source_row, c).value, "date")
    ]
    for idx, col in enumerate(date_cols[:5]):
        ws.cell(target_row, col).value = _fecha_excel(lunes + timedelta(days=idx))

    rotacion = cfg["rotacion_pmp"]
    set_rot = set(rotacion)
    set_c3 = set(cfg.get("clientes_celula3", []))
    # Columnas "Consultor" del bloque (tabla izquierda y tabla-resumen derecha;
    # se detectan por el encabezado, así aguantan que se hayan movido de columna).
    cols_cons = {c for c in range(1, ws.max_column + 1)
                 if str(ws.cell(target_row, c).value or "").strip() == "Consultor"}

    # Rotación: por cada fila de datos, si atiende un cliente de Célula 3 se pone el
    # SIGUIENTE encargado (giro de cfg['rotacion_pmp']) en sus columnas de consultor;
    # si el cliente NO es de Célula 3, el consultor se deja en blanco (este documento
    # es solo de C3). Todo lo demás (clientes, horarios, formato) queda idéntico.
    for row in range(target_row + 1, target_row + alto):
        cli_c3 = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row, c).value
            if isinstance(v, str) and v.strip() in set_c3:
                cli_c3 = v.strip(); break
        if cli_c3:
            actual = None                       # encargado de C3 más a la derecha (tabla resumen)
            for c in range(ws.max_column, 0, -1):
                v = ws.cell(row, c).value
                if isinstance(v, str) and v.strip() in set_rot:
                    actual = v.strip(); break
            if actual:
                nuevo = siguiente_disponible(actual, rotacion, ausentes, cobertura)
                for c in cols_cons:
                    ws.cell(row, c).value = nuevo
        else:
            for c in cols_cons:
                v = ws.cell(row, c).value
                if isinstance(v, str) and v.strip() and v.strip() != "Consultor":
                    ws.cell(row, c).value = None

    # FESTIVO por día: para CADA día hábil de la semana, la columna debe decir
    # "FESTIVO" (celda combinada vertical) si la fecha es festiva en Colombia, y
    # mostrar los clientes si no. El bloque se copió del anterior, así que cada
    # columna puede venir como FESTIVO o con clientes según la semana origen → se
    # ajusta en AMBOS sentidos. (Antes solo se trataba —y solo para quitar— el lunes,
    # por lo que un día festivo entre semana, o un lunes festivo que el origen no
    # traía marcado, quedaban con el cliente en vez de "FESTIVO".)
    if date_cols:
        ref_festivo = _celda_festivo_ref(ws)
        fila_ini, fila_fin = _rango_datos_bloque(ws, target_row, alto, date_cols)
        for idx, col in enumerate(date_cols[:5]):
            dia = lunes + timedelta(days=idx)
            merge = _merge_festivo_en_col(ws, col, target_row, target_row + alto)
            if es_festivo(dia):
                if merge is None:
                    _poner_festivo_col(ws, col, fila_ini, fila_fin, ref_festivo)
            elif merge is not None:
                _quitar_festivo_col(ws, col, merge, cols_cons)

    # Ancho de columna: al moverse la tabla-resumen de C3 a la columna L (más
    # estrecha que la M que usaba antes), los nombres salían cortados. Se igualan
    # las columnas de "Consultor" al ancho de la más ancha para que se lean enteros.
    from openpyxl.utils import get_column_letter
    anchos = [ws.column_dimensions[get_column_letter(c)].width or 0 for c in cols_cons]
    if anchos:
        ancho_obj = max(anchos)
        for c in cols_cons:
            letra = get_column_letter(c)
            if (ws.column_dimensions[letra].width or 0) < ancho_obj:
                ws.column_dimensions[letra].width = ancho_obj

    destino = salida or str(Path(ruta_pmp).with_name(
        f"{Path(ruta_pmp).stem}_actualizado_{lunes.strftime('%Y%m%d')}.xlsx"
    ))
    _guardar_wb(wb, destino)
    return destino


def _ultima_fila_dispo(ws) -> Optional[int]:
    ultima = None
    for row in range(1, ws.max_row + 1):
        if hasattr(ws.cell(row, 2).value, "date") and hasattr(ws.cell(row, 3).value, "date"):
            ultima = row
    return ultima


def cobertura_dispo(ruta_matriz: str) -> Optional[date]:
    """Último viernes (fin de disponibilidad) cubierto en la Matriz. Sirve para
    avisar al usuario hasta cuándo está rellena y cuándo tocará extenderla."""
    from openpyxl import load_workbook
    try:
        ws = load_workbook(ruta_matriz, data_only=True)["ROTACION DISPO CELULA 3 "]
    except Exception:
        return None
    fin = None
    for row in range(1, ws.max_row + 1):
        c = ws.cell(row, 3).value
        if hasattr(c, "date") and (fin is None or c.date() > fin):
            fin = c.date()
    return fin


def asegurar_dispo_en_matriz(ruta_matriz: str, inicio: date, cfg: dict,
                             salida: Optional[str] = None) -> str:
    """Asegura que la matriz tenga la fila V-V de `inicio`.

    Si ya existe, solo guarda una copia equivalente. Si falta, agrega filas
    copiando el formato de la última fila real y rotando N2/N3.
    """
    from openpyxl import load_workbook

    wb = load_workbook(ruta_matriz)
    ws = wb["ROTACION DISPO CELULA 3 "]
    ultima = _ultima_fila_dispo(ws)
    if not ultima:
        raise ValueError("No encontré filas de disponibilidad en la matriz.")

    for row in range(1, ws.max_row + 1):
        valor = ws.cell(row, 2).value
        if hasattr(valor, "date") and valor.date() == inicio:
            destino = salida or str(Path(ruta_matriz).with_name(
                f"{Path(ruta_matriz).stem}_actualizada_{inicio.strftime('%Y%m%d')}.xlsx"
            ))
            _guardar_wb(wb, destino)
            return destino

    rotacion_n3 = cfg.get("rotacion_n3") or []
    if not rotacion_n3:
        rotacion_n3 = []
        for row in range(1, ws.max_row + 1):
            nombre = ws.cell(row, 5).value
            if nombre and nombre not in rotacion_n3:
                rotacion_n3.append(str(nombre))

    while ws.cell(ultima, 2).value.date() < inicio:
        nueva = ultima + 1
        for col in range(2, 6):
            origen = ws.cell(ultima, col)
            destino = ws.cell(nueva, col)
            if origen.has_style:
                destino._style = copy.copy(origen._style)
            destino.number_format = origen.number_format
            destino.font = copy.copy(origen.font)
            destino.fill = copy.copy(origen.fill)
            destino.border = copy.copy(origen.border)
            destino.alignment = copy.copy(origen.alignment)
            destino.protection = copy.copy(origen.protection)
        ws.row_dimensions[nueva].height = ws.row_dimensions[ultima].height

        inicio_nuevo = ws.cell(ultima, 3).value.date()
        ws.cell(nueva, 2).value = _fecha_excel(inicio_nuevo)
        ws.cell(nueva, 3).value = _fecha_excel(inicio_nuevo + timedelta(days=7))
        ws.cell(nueva, 4).value = _siguiente_en_rotacion(str(ws.cell(ultima, 4).value), cfg["rotacion_n2"])
        ws.cell(nueva, 5).value = _siguiente_en_rotacion(str(ws.cell(ultima, 5).value), rotacion_n3)
        ultima = nueva

    destino = salida or str(Path(ruta_matriz).with_name(
        f"{Path(ruta_matriz).stem}_actualizada_{inicio.strftime('%Y%m%d')}.xlsx"
    ))
    _guardar_wb(wb, destino)
    return destino


# ══════════════════════════════════════════════════════════════════════════════
#  4 · MODOS
# ══════════════════════════════════════════════════════════════════════════════
def modo_dispo(cfg: dict) -> None:
    console.print(Panel("🌙  Consulta de disponibilidad nocturna",
                        border_style="magenta", box=box.ROUNDED, expand=False))
    lunes = pedir_fecha("Semana a consultar (lunes)", siguiente_lunes())

    n2 = n3 = None
    fuente = "cálculo automático"
    ruta_matriz = resolver_ruta(cfg, "ruta_matriz", "Matriz", "Matriz Unificada", requerido=False)
    if ruta_matriz:
        with console.status("[cyan]Leyendo Matriz Unificada…", spinner="dots"):
            try:
                n2, n3 = leer_dispo(lunes, ruta_matriz)
                fuente = "Matriz Unificada"
            except Exception as e:
                console.print(f"  [yellow]⚠[/] {e}")
    if not n2:
        n2 = calcular_dispo_fallback(lunes, cfg)

    fin = (lunes + timedelta(4)).strftime("%d/%m/%Y")
    cuerpo = Text()
    cuerpo.append("N2 (incidentes nocturnos) : ", style="bold"); cuerpo.append(f"{n2}\n", style="cyan")
    cuerpo.append("N3 (escalamiento)         : ", style="bold"); cuerpo.append(f"{n3 or 'N/A'}\n")
    cuerpo.append("Fuente                    : ", style="bold"); cuerpo.append(fuente, style="dim")
    if ruta_matriz:
        cob = cobertura_dispo(ruta_matriz)
        if cob:
            cuerpo.append("\nCubierta hasta            : ", style="bold")
            cuerpo.append(cob.strftime("%d/%m/%Y"), style="dim")
    console.print(Panel(cuerpo, title=f"🌙  Disponibilidad — {lunes.strftime('%d/%m/%Y')} → {fin}",
                        border_style="yellow", box=box.ROUNDED, expand=False, padding=(1, 2)))


def modo_pmp(cfg: dict) -> None:
    console.print(Panel("📋  Generar semana siguiente\n"
                        "[dim]Actualiza el Control PMP y la Matriz, y crea el cuadro resumen.[/]",
                        border_style="red", box=box.ROUNDED, expand=False))

    # 1 · Resolver archivos primero (para poder sugerir la fecha correcta del Control).
    #     Antes de nada, preferir SIEMPRE el Control con la semana más avanzada: tras
    #     generar una semana, su salida («_actualizado_AAAAMMDD.xlsx») queda junto a la
    #     fuente; sin esto se seguiría partiendo del archivo viejo y se propondría una
    #     semana ya hecha. Así el flujo encadena solo.
    mejor = control_mas_reciente(cfg)
    if mejor:
        prev = cfg.get("ruta_pmp", "")
        prev_resuelto = str(Path(prev).resolve()) if prev and Path(prev).is_file() else ""
        if mejor != prev_resuelto:
            cfg["ruta_pmp"] = mejor
            guardar_cfg(cfg)
            if prev_resuelto:   # solo avisar si REEMPLAZA a un Control previo distinto
                console.print(f"  [green]✓[/] Usando el Control más reciente: "
                              f"[cyan]{_etiqueta_archivo(Path(mejor))}[/]")
    ruta_pmp = resolver_ruta(cfg, "ruta_pmp", "PMP", "Control de Gestión PMP", requerido=True)
    if not ruta_pmp:
        console.print("  [red]✗[/] Sin PMP no se puede generar."); return
    ruta_matriz = resolver_ruta(cfg, "ruta_matriz", "Matriz", "Matriz Unificada", requerido=False)

    # 2 · Fecha AUTOMÁTICA: se detecta la última semana registrada en el Control y se
    #     usa la siguiente. El usuario no escribe fechas; solo confirma. Si necesita
    #     otra (caso raro), puede escribirla al responder que no.
    try:
        sugerida = siguiente_lunes_en_pmp(ruta_pmp) or siguiente_lunes()
    except Exception:
        sugerida = siguiente_lunes()
    fin_sug = (sugerida + timedelta(4)).strftime("%d/%m/%Y")
    console.print(f"  [green]✓[/] Detectada la última semana del Control. Se generará la "
                  f"siguiente: [cyan]{sugerida.strftime('%d/%m/%Y')} → {fin_sug}[/].")
    if confirmar(f"¿Generar la semana del {sugerida.strftime('%d/%m/%Y')}?", default=True):
        lunes_sig = sugerida
    else:
        lunes_sig = pedir_fecha("Semana a GENERAR (lunes)", sugerida)

    # 2b · Validar coherencia AHORA, antes de leer / clasificar / previsualizar.
    ok, msg, sug = diagnostico_fecha(ruta_pmp, lunes_sig)
    if not ok:
        console.print(msg)
        if sug and confirmar(f"¿Usar {sug.strftime('%d/%m/%Y')} en su lugar?", default=True):
            lunes_sig = sug
        else:
            console.print("  [dim]Cancelado — no se escribió ningún archivo.[/]"); return
    lunes_act = lunes_sig - timedelta(weeks=1)

    # 3 · Ausentes + cobertura (quién cubre a cada ausente, lo elige el usuario)
    ausentes = seleccionar_ausentes(cfg["rotacion_pmp"])
    cobertura = pedir_cobertura(cfg["rotacion_pmp"], ausentes) if ausentes else {}

    # 4 · Leer el reparto actual de Célula 3 — por NOMBRE de cliente, robusto a que
    #     las columnas se hayan movido (tabla resumen en M o en L).
    with console.status("[cyan]Leyendo el reparto actual de Célula 3…", spinner="dots"):
        actuales, faltantes = leer_reparto_actual(ruta_pmp, lunes_act, cfg)
    if all(not v for v in actuales.values()):
        console.print(f"  [yellow]⚠[/] No se halló el reparto de Célula 3 para la semana "
                      f"{lunes_act.strftime('%d/%m/%Y')} en el Control."); return
    if faltantes:
        console.print(f"  [yellow]⚠[/] Clientes no ubicados esa semana: "
                      f"{', '.join(faltantes)} — revísalo en la vista previa.")

    # 4c · Rotar (regla determinista) + disponibilidad
    nuevas = rotar(actuales, ausentes, cfg["rotacion_pmp"], cobertura)
    n2 = n3 = None
    fuente_dispo = "cálculo automático"
    if ruta_matriz:
        try:
            n2, n3 = leer_dispo(lunes_sig, ruta_matriz)
            if n2:
                fuente_dispo = "Matriz Unificada"
        except Exception as e:
            console.print(f"  [yellow]⚠[/] Dispo: {e}")
    if not n2:
        n2 = calcular_dispo_fallback(lunes_sig, cfg)

    # Mostrar de dónde salió cada cliente (semana de referencia)
    ref = Text(f"Semana de referencia leída: {lunes_act.strftime('%d/%m/%Y')}\n", style="dim")
    for ing, cls in actuales.items():
        if cls:
            ref.append(f"  {ing}: ", style="dim")
            ref.append(", ".join(cls) + "\n")
    console.print(ref)

    # 5 · PREVIEW + confirmación
    console.print(tabla_preview(lunes_sig, nuevas, ausentes, n2, n3, fuente_dispo, cfg))
    if not confirmar("¿Confirmar y generar la semana siguiente?", default=True):
        console.print("  [dim]Cancelado — no se escribió ningún archivo.[/]"); return

    # 6 · A partir de la MISMA rotación: Control real + Matriz + cuadro resumen.
    #     El Control y la Matriz se SOBRESCRIBEN sobre su mismo archivo (in-place, sin
    #     generar copias nuevas): se pasa salida=ruta_* y el guardado es atómico
    #     (temp + os.replace), así un fallo a mitad no corrompe el archivo. El Control
    #     es la operación con validaciones: si falla, se aborta sin tocar nada.
    dir_destino = Path(ruta_pmp).resolve().parent
    try:
        with console.status("[green]Actualizando el Control PMP…", spinner="dots"):
            control = actualizar_control_pmp(ruta_pmp, lunes_sig, ausentes, cfg, salida=ruta_pmp,
                                             cobertura=cobertura)
    except Exception as e:
        console.print(f"  [red]✗[/] No se pudo actualizar el Control PMP: {e}")
        return

    matriz = None
    if ruta_matriz:
        try:
            with console.status("[green]Actualizando la Matriz de recursos…", spinner="dots"):
                matriz = asegurar_dispo_en_matriz(ruta_matriz, lunes_sig - timedelta(days=3), cfg,
                                                  salida=ruta_matriz)
            cob = cobertura_dispo(matriz or ruta_matriz)
            if cob:
                console.print(f"  [dim]🌙 Disponibilidad nocturna cubierta hasta el "
                              f"{cob.strftime('%d/%m/%Y')}.[/]")
        except Exception as e:
            console.print(f"  [yellow]⚠[/] No se pudo actualizar la Matriz (se omite): {e}")

    # Cuadro resumen: tabla NUEVA y formateada, aparte, para compartir al equipo.
    # Es opcional (por defecto no): el entregable principal es el Control copiado+rotado.
    resumen = None
    if confirmar("¿Generar además un cuadro resumen aparte para compartir al equipo?",
                 default=False):
        try:
            with console.status("[green]Generando el cuadro resumen…", spinner="dots"):
                resumen = generar_excel(lunes_sig, nuevas, n2, n3, ausentes, cfg, dir_destino)
        except Exception as e:
            console.print(f"  [red]✗[/] Error al generar el cuadro resumen: {e}")
            resumen = None

    # 7 · Resultado
    fin = (lunes_sig + timedelta(4)).strftime("%d/%m/%Y")
    res = Text()
    res.append("Semana        : ", style="bold"); res.append(f"{lunes_sig.strftime('%d/%m/%Y')} → {fin}\n")
    res.append("Dispo N2      : ", style="bold"); res.append(f"{n2}\n", style="yellow")
    res.append("Dispo N3      : ", style="bold"); res.append(f"{n3 or 'N/A'}\n\n")
    res.append("Control PMP   : ", style="bold"); res.append(f"{control}\n", style="cyan")
    res.append("Matriz        : ", style="bold"); res.append(f"{matriz or 'no actualizada'}\n",
                                                              style="cyan" if matriz else "dim")
    res.append("Cuadro resumen: ", style="bold"); res.append(f"{resumen or 'no generado'}",
                                                              style="cyan" if resumen else "dim")
    console.print(Panel(res, title="✅  Semana siguiente generada",
                        border_style="green", box=box.DOUBLE, expand=False, padding=(1, 2)))
    console.print("  [dim]Siguiente paso: revisa el Control PMP actualizado y súbelo a "
                  "SharePoint. El cuadro resumen es para compartirlo al equipo.[/]\n")

    if confirmar("¿Abrir el Control PMP actualizado ahora?", default=True):
        abrir_archivo(control)
    if resumen and confirmar("¿Abrir también el cuadro resumen?", default=False):
        abrir_archivo(resumen)


def _config_rutas(cfg: dict) -> None:
    console.print(Panel(f"📁  Rutas de archivos\n[dim]Se guardan en {CONFIG_PATH}[/]",
                        border_style="blue", box=box.ROUNDED, expand=False))

    def _editar(clave: str, desc: str, patron: str, requerido: bool) -> None:
        actual = cfg.get(clave, "")
        estado = f"[cyan]{_etiqueta_archivo(Path(actual))}[/]" if actual and Path(actual).is_file() \
            else ("[yellow](sin configurar)[/]" if not actual else "[red](ruta inválida)[/]")
        console.print(f"  {desc}: {estado}")
        accion = _ask(questionary.select(
            f"{desc} — ¿qué hacer?",
            choices=[questionary.Choice("Auto-detectar / volver a buscar", "auto"),
                     questionary.Choice("Escribir ruta manual", "manual"),
                     questionary.Choice("Dejar como está", "skip")],
            style=QS))
        if accion == "skip":
            return
        if accion == "auto":
            cfg[clave] = ""                       # forzar re-detección
            resolver_ruta(cfg, clave, patron, desc, requerido=requerido)
        else:
            _pedir_ruta_manual(cfg, clave, desc, requerido)

    _editar("ruta_pmp", "Control de Gestión PMP", "PMP", True)
    _editar("ruta_matriz", "Matriz Unificada", "Matriz", False)
    guardar_cfg(cfg)
    console.print("  [green]✓[/] Rutas guardadas.")


def _pedir_nombre(prompt: str) -> str:
    """Pide un nombre y normaliza espacios. Cadena vacía = cancelar/sin cambios."""
    return " ".join(_ask(questionary.text(prompt, style=QS)).split())


# Sentinela para las opciones "cancelar" de los select de abajo.
_CANCELAR = "↩  cancelar"


def _gestionar_personas(cfg: dict, clave: str, titulo: str, es_giro: bool) -> None:
    """Añadir / quitar / reordenar una lista de consultores (rotación PMP, N2 o N3).
    El ORDEN importa: en PMP define el giro (cada cliente pasa al SIGUIENTE de la
    lista); en N2/N3 define la secuencia de turnos. Solo edita la config; los
    nombres deben coincidir EXACTAMENTE con los del Excel para que la lectura y la
    rotación funcionen."""
    nota = None
    while True:
        banner()                                   # limpia la pantalla: no se apilan paneles
        lista: List[str] = cfg.setdefault(clave, [])
        cuerpo = Text()
        if lista:
            for i, n in enumerate(lista, 1):
                cuerpo.append(f"  {i}. {n}\n")
            if es_giro and len(lista) >= 2:
                cuerpo.append("\n  Giro: " + " → ".join(lista) + " → " + lista[0], style="dim")
        else:
            cuerpo.append("  (lista vacía)", style="yellow")
        console.print(Panel(cuerpo, title=f"👥  {titulo}", border_style="blue",
                            box=box.ROUNDED, expand=False, padding=(1, 2)))
        if nota:
            console.print(nota); nota = None
        accion = _ask(questionary.select(
            f"{titulo} — ¿qué hacer?",
            choices=[questionary.Choice("➕  Añadir consultor", "add"),
                     questionary.Choice("➖  Quitar consultor", "del"),
                     questionary.Choice("↕   Reordenar", "move"),
                     questionary.Choice("↩   Volver", "back")],
            style=QS))
        if accion == "back":
            return
        if accion == "add":
            nombre = _pedir_nombre("Nombre del consultor (TAL CUAL aparece en el Excel):")
            if not nombre:
                nota = "  [yellow]·[/] Sin cambios."; continue
            if nombre in lista:
                nota = f"  [yellow]·[/] «{nombre}» ya está en la lista."; continue
            lista.append(nombre); guardar_cfg(cfg)
            nota = f"  [green]✓[/] Añadido «{nombre}» al final. Usa «Reordenar» si va en otra posición."
        elif accion == "del":
            if not lista:
                nota = "  [yellow]·[/] La lista ya está vacía."; continue
            quien = _ask(questionary.select("¿A quién quitar?", choices=lista + [_CANCELAR], style=QS))
            if quien == _CANCELAR:
                continue
            lista.remove(quien); guardar_cfg(cfg)
            nota = f"  [green]✓[/] Quitado «{quien}»."
            if es_giro and len(lista) < 2:
                nota += "\n  [yellow]⚠[/] Con menos de 2 consultores el giro no rota."
        elif accion == "move":
            if len(lista) < 2:
                nota = "  [yellow]·[/] Hacen falta al menos 2 para reordenar."; continue
            quien = _ask(questionary.select("¿Cuál mover?", choices=lista + [_CANCELAR], style=QS))
            if quien == _CANCELAR:
                continue
            otras = [c for c in lista if c != quien]
            destinos = [questionary.Choice("al principio", 0)] + \
                       [questionary.Choice(f"después de {c}", i + 1) for i, c in enumerate(otras)]
            pos = _ask(questionary.select(f"¿Dónde colocar a «{quien}»?", choices=destinos, style=QS))
            lista.remove(quien); lista.insert(int(pos), quien); guardar_cfg(cfg)
            nota = "  [green]✓[/] Nuevo orden: " + " → ".join(lista)


def _set_atributos_cliente(cfg: dict, cliente: str) -> None:
    """Pregunta horario (mañana/tarde) y si es PMP largo, y sincroniza las listas
    horario_tarde / clientes_largos para `cliente` (usadas en el cuadro resumen)."""
    horario = _ask(questionary.select(
        f"Horario de «{cliente}»:",
        choices=[questionary.Choice("Mañana (7:00–12:00)", "manana"),
                 questionary.Choice("Tarde (12:00–17:00)", "tarde")],
        style=QS))
    largo = confirmar(f"¿«{cliente}» es PMP largo?",
                      default=cliente in set(cfg.get("clientes_largos", [])))
    tarde = [c for c in cfg.get("horario_tarde", []) if c != cliente]
    if horario == "tarde":
        tarde.append(cliente)
    cfg["horario_tarde"] = tarde
    largos = [c for c in cfg.get("clientes_largos", []) if c != cliente]
    if largo:
        largos.append(cliente)
    cfg["clientes_largos"] = largos


def _gestionar_clientes(cfg: dict) -> None:
    """Añadir / quitar clientes de Célula 3 y editar sus atributos (horario, PMP
    largo). Solo edita la config: el cliente debe existir ya en el Excel con el
    MISMO nombre para que la herramienta lo lea y lo rote."""
    nota = None
    while True:
        banner()                                   # limpia la pantalla: no se apilan tablas
        clientes: List[str] = cfg.setdefault("clientes_celula3", [])
        tarde = set(cfg.get("horario_tarde", []))
        largos = set(cfg.get("clientes_largos", []))
        tabla = Table(box=box.SIMPLE, expand=False, title="🧾  Clientes de Célula 3")
        tabla.add_column("#", justify="right", style="dim")
        tabla.add_column("Cliente")
        tabla.add_column("Horario")
        tabla.add_column("PMP largo", justify="center")
        if clientes:
            for i, c in enumerate(clientes, 1):
                tabla.add_row(str(i), c, "Tarde" if c in tarde else "Mañana",
                              "Sí" if c in largos else "—")
        else:
            tabla.add_row("", "[yellow](sin clientes)[/]", "", "")
        console.print(tabla)
        if nota:
            console.print(nota); nota = None
        accion = _ask(questionary.select(
            "Clientes de Célula 3 — ¿qué hacer?",
            choices=[questionary.Choice("➕  Añadir cliente", "add"),
                     questionary.Choice("➖  Quitar cliente", "del"),
                     questionary.Choice("✎   Editar horario / PMP largo", "edit"),
                     questionary.Choice("↩   Volver", "back")],
            style=QS))
        if accion == "back":
            return
        if accion == "add":
            nombre = _pedir_nombre("Nombre del cliente (TAL CUAL aparece en el Excel):")
            if not nombre:
                nota = "  [yellow]·[/] Sin cambios."; continue
            if nombre in clientes:
                nota = f"  [yellow]·[/] «{nombre}» ya está en la lista."; continue
            clientes.append(nombre)
            _set_atributos_cliente(cfg, nombre)
            guardar_cfg(cfg)
            nota = f"  [green]✓[/] Añadido «{nombre}»."
        elif accion in ("del", "edit"):
            if not clientes:
                nota = "  [yellow]·[/] No hay clientes."; continue
            verbo = "quitar" if accion == "del" else "editar"
            quien = _ask(questionary.select(f"¿Qué cliente {verbo}?",
                                            choices=clientes + [_CANCELAR], style=QS))
            if quien == _CANCELAR:
                continue
            if accion == "del":
                clientes.remove(quien)
                cfg["horario_tarde"] = [c for c in cfg.get("horario_tarde", []) if c != quien]
                cfg["clientes_largos"] = [c for c in cfg.get("clientes_largos", []) if c != quien]
                guardar_cfg(cfg)
                nota = f"  [green]✓[/] Quitado «{quien}» (y sus atributos)."
            else:
                _set_atributos_cliente(cfg, quien)
                guardar_cfg(cfg)
                nota = f"  [green]✓[/] Atributos de «{quien}» actualizados."


def modo_config(cfg: dict) -> None:
    while True:
        banner()                                   # limpia la pantalla en cada vuelta
        console.print(Panel(f"⚙  Configuración\n[dim]Se guarda en {CONFIG_PATH}[/]",
                            border_style="blue", box=box.ROUNDED, expand=False))
        op = _ask(questionary.select(
            "¿Qué quieres configurar?",
            choices=[
                questionary.Choice("📁  Rutas de archivos (Control / Matriz)", "rutas"),
                questionary.Choice("👥  Consultores — rotación PMP (Célula 3)", "pmp"),
                questionary.Choice("🌙  Consultores — N2 (incidentes nocturnos)", "n2"),
                questionary.Choice("🌙  Consultores — N3 (escalamiento)", "n3"),
                questionary.Choice("🧾  Clientes de Célula 3 (rotación + horario)", "cli"),
                questionary.Choice("↩   Volver al menú", "back"),
            ],
            style=QS))
        if op == "back":
            return
        if op == "rutas":
            _config_rutas(cfg)
        elif op == "pmp":
            _gestionar_personas(cfg, "rotacion_pmp", "Rotación PMP (Célula 3)", es_giro=True)
        elif op == "n2":
            _gestionar_personas(cfg, "rotacion_n2", "Rotación N2 (incidentes nocturnos)", es_giro=True)
        elif op == "n3":
            _gestionar_personas(cfg, "rotacion_n3", "Rotación N3 (escalamiento)", es_giro=True)
        elif op == "cli":
            _gestionar_clientes(cfg)


# ══════════════════════════════════════════════════════════════════════════════
#  5 · MENÚ PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def menu(cfg: dict) -> None:
    auto_detectar_inicio(cfg)
    primera = True
    while True:
        if not primera:
            banner()
        primera = False

        rutas_ok = bool(cfg.get("ruta_pmp") and Path(cfg["ruta_pmp"]).is_file())
        estado = "[green]rutas OK ✓[/]" if rutas_ok else "[yellow]configura rutas ⚠[/]"
        console.print(f"  Estado: {estado}\n")

        try:
            op = questionary.select(
                "¿Qué deseas hacer?",
                choices=[
                    questionary.Choice("🌙  Consultar disponibilidad nocturna", "1"),
                    questionary.Choice("📋  Generar semana siguiente", "2"),
                    questionary.Choice("⚙   Configurar (rutas · consultores · clientes)", "3"),
                    questionary.Choice("🚪  Salir", "0"),
                ],
                style=QS, qmark="›", pointer="❯").unsafe_ask()
        except (KeyboardInterrupt, EOFError):
            break

        if op in (None, "0"):
            break

        accion = {"1": modo_dispo, "2": modo_pmp, "3": modo_config}.get(op)
        if accion:
            try:
                console.print()
                accion(cfg)
            except Volver:
                console.print("  [dim]↩  Volviendo al menú…[/]")
            except Exception as e:
                # Red de seguridad: nunca un traceback crudo delante del público.
                console.print(Panel(f"[red]{type(e).__name__}:[/] {e}",
                                    title="⚠  Ocurrió un error inesperado",
                                    border_style="red", box=box.ROUNDED, expand=False,
                                    padding=(0, 2)))
            try:
                questionary.text("", qmark="",
                                 instruction="(Enter para volver al menú)").unsafe_ask()
            except (KeyboardInterrupt, EOFError):
                break
        banner()

    limpiar()
    console.print("\n  [dim]Hasta luego 👋[/]\n")


def _selftest() -> None:
    import tempfile
    from openpyxl import Workbook, load_workbook

    cfg = DEFAULT_CFG.copy()
    assert parse_fecha("19/06/2026") == date(2026, 6, 19)
    assert parse_fecha("2026-06-19") == date(2026, 6, 19)
    actuales = {
        "Mateo Florez": ["A"],
        "Heiner Diaz": ["B"],
        "Estefania Sanabria": ["C"],
    }
    # Giro Heiner→Mateo→Estefania→Heiner. Con Heiner ausente, su destino salta.
    rotadas = rotar(actuales, ["Heiner Diaz"], cfg["rotacion_pmp"])
    assert rotadas["Estefania Sanabria"] == ["A"]       # Mateo→Estefania
    assert rotadas["Mateo Florez"] == ["B", "C"]         # Heiner→Mateo; Estefania→(Heiner ausente)→Mateo
    assert rotadas["Heiner Diaz"] == []

    # siguiente_disponible: la regla única de rotación, sin y con ausentes.
    rot = cfg["rotacion_pmp"]
    assert siguiente_disponible("Heiner Diaz", rot, []) == "Mateo Florez"
    assert siguiente_disponible("Mateo Florez", rot, []) == "Estefania Sanabria"
    assert siguiente_disponible("Estefania Sanabria", rot, []) == "Heiner Diaz"
    assert siguiente_disponible("Estefania Sanabria", rot, ["Heiner Diaz"]) == "Mateo Florez"
    # Cobertura explícita: el destino que sería un ausente lo toma quien diga la
    # cobertura (elección del usuario), NO el salto automático al siguiente.
    assert siguiente_disponible("Estefania Sanabria", rot, ["Heiner Diaz"],
                                {"Heiner Diaz": "Estefania Sanabria"}) == "Estefania Sanabria"
    rc = rotar({"Mateo Florez": ["A"], "Heiner Diaz": ["B"], "Estefania Sanabria": ["C"]},
               ["Heiner Diaz"], rot, {"Heiner Diaz": "Estefania Sanabria"})
    assert rc["Mateo Florez"] == ["B"] and rc["Estefania Sanabria"] == ["A", "C"] \
        and rc["Heiner Diaz"] == [], rc

    # Festivos de Colombia (deterministas): los que de hecho aparecen en el Control.
    assert es_festivo(date(2026, 6, 8))     # Corpus Christi (trasladado)
    assert es_festivo(date(2026, 6, 15))    # Sagrado Corazón
    assert not es_festivo(date(2026, 6, 22))
    assert es_festivo(date(2026, 6, 29))    # San Pedro y Pablo (trasladado)

    with tempfile.TemporaryDirectory() as tmp:
        pmp = Path(tmp) / "pmp.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "2026"
        ws["B1"] = "Consultor"
        for col, d in zip((4, 6, 8, 9, 15), [date(2026, 6, 22) + timedelta(days=i) for i in range(5)]):
            ws.cell(1, col).value = _fecha_excel(d)
        ws["B3"] = "Mateo Florez"
        wb.save(pmp)

        # Sugerencia / validación de fecha (determinista): el bloque del 22/06 ya
        # existe, así que la siguiente que falta es el 29/06.
        assert siguiente_lunes_en_pmp(str(pmp)) == date(2026, 6, 29)
        assert diagnostico_fecha(str(pmp), date(2026, 6, 22))[0] is False   # ya existe
        assert diagnostico_fecha(str(pmp), date(2026, 6, 29))[0] is True    # la que falta

        matriz = Path(tmp) / "matriz.xlsx"
        matriz_out = Path(tmp) / "matriz_out.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "ROTACION DISPO CELULA 3 "
        ws["B4"] = _fecha_excel(date(2026, 6, 19))
        ws["C4"] = _fecha_excel(date(2026, 6, 26))
        ws["D4"] = "Heiner Diaz"
        ws["E4"] = "Carlos Barrera"
        wb.save(matriz)
        asegurar_dispo_en_matriz(str(matriz), date(2026, 6, 26), cfg, str(matriz_out))
        wb = load_workbook(matriz_out, data_only=False)
        ws = wb["ROTACION DISPO CELULA 3 "]
        assert ws["B5"].value.date() == date(2026, 6, 26)
        assert ws["D5"].value == "Mateo Florez"
        assert cobertura_dispo(matriz_out) == date(2026, 7, 3)

    # ── Estructura NUEVA del Control: tabla-resumen de C3 en la columna L, días en
    #    E/F/H/I/N, dos tablas lado a lado. Verifica lectura por NOMBRE y rotación en
    #    AMBAS tablas, inmune a que la columna del consultor se haya movido. ──
    with tempfile.TemporaryDirectory() as tmp2:
        ruta = Path(tmp2) / "control_nuevo.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "2026"

        def _bloque(fila, lunes_b, filas_datos):
            ws.cell(fila, 2).value = "Consultor"          # tabla izquierda
            ws.cell(fila, 12).value = "Consultor"         # tabla resumen (derecha, col L)
            for col, d in zip((4, 6, 8, 9, 14),
                              [lunes_b + timedelta(days=i) for i in range(5)]):
                ws.cell(fila, col).value = _fecha_excel(d)
            for i, (b_izq, cliente, l_der) in enumerate(filas_datos, start=1):
                r = fila + i
                ws.cell(r, 2).value = b_izq               # consultor operación (izq)
                ws.cell(r, 5).value = cliente             # cliente por día (col E)
                ws.cell(r, 12).value = l_der              # consultor C3 (derecha, col L)
                ws.cell(r, 14).value = cliente            # cliente (tabla resumen)

        _bloque(1, date(2026, 6, 15), [("Heiner Diaz", "HOMI Oracle", "Heiner Diaz")])
        # 'La Riviera (APP)' la lleva en la izquierda alguien de FUERA de C3 (Susana),
        # pero la tabla resumen (derecha) la asigna a Estefania → manda la derecha.
        # 'D1 Oracle' no es de C3 → su consultor debe quedar en blanco al copiar.
        _bloque(13, date(2026, 6, 22), [
            ("Mateo Florez",     "HOMI Oracle",        "Mateo Florez"),
            ("Susana Rodriguez", "La Riviera (APP)",   "Estefania Sanabria"),
            ("Heiner Diaz",      "Bancoldex (Oracle)", "Heiner Diaz"),
            ("Juan Pablo Lombo", "D1 Oracle",          "Juan Pablo Lombo"),
        ])
        wb.save(ruta)

        rep, _falt = leer_reparto_c3(load_workbook(ruta, data_only=True)["2026"],
                                     date(2026, 6, 22), cfg["clientes_celula3"], cfg["rotacion_pmp"])
        assert rep["Mateo Florez"] == ["HOMI Oracle"], rep
        assert rep["Estefania Sanabria"] == ["La Riviera (APP)"], rep
        assert rep["Heiner Diaz"] == ["Bancoldex (Oracle)"], rep
        assert _espaciado_bloques(_headers_semana_pmp(load_workbook(ruta)["2026"])) == 12

        out = Path(tmp2) / "control_out.xlsx"
        actualizar_control_pmp(str(ruta), date(2026, 6, 29), [], cfg, str(out))
        ws2 = load_workbook(out, data_only=False)["2026"]
        assert ws2.cell(25, 4).value.date() == date(2026, 6, 29)     # D = lunes nuevo
        assert ws2.cell(25, 14).value.date() == date(2026, 7, 3)     # N = viernes nuevo
        # HOMI Oracle (C3): Mateo→Estefania en AMBAS tablas; el cliente intacto.
        assert ws2.cell(26, 2).value == "Estefania Sanabria"
        assert ws2.cell(26, 12).value == "Estefania Sanabria"
        assert ws2.cell(26, 5).value == "HOMI Oracle"
        # La Riviera (C3): Estefania→Heiner en ambas tablas; la izquierda, que tenía a
        # Susana, también toma el encargado de C3 porque el cliente SÍ es de C3.
        assert ws2.cell(27, 2).value == "Heiner Diaz"
        assert ws2.cell(27, 12).value == "Heiner Diaz"
        # Bancoldex (C3): Heiner→Mateo.
        assert ws2.cell(28, 2).value == "Mateo Florez"
        assert ws2.cell(28, 12).value == "Mateo Florez"
        # D1 Oracle (NO es de C3): el consultor queda en blanco en ambas tablas; el
        # cliente se conserva.
        assert ws2.cell(29, 2).value is None
        assert ws2.cell(29, 12).value is None
        assert ws2.cell(29, 5).value == "D1 Oracle"
        # Lunes 29/06/2026 es festivo (San Pedro y San Pablo) → su columna = FESTIVO.
        assert ws2.cell(26, 4).value == "FESTIVO", ws2.cell(26, 4).value

    # ── Festivos en CUALQUIER día hábil: la semana del 30/03/2026 tiene Jueves
    #    (02/04) y Viernes (03/04) Santo festivos; lun/mar/mié no. Cada día festivo
    #    debe quedar "FESTIVO" (combinado) y los no festivos conservar su cliente.
    with tempfile.TemporaryDirectory() as tmp3:
        ruta = Path(tmp3) / "control_fest.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "2026"
        cols_dia = (4, 6, 8, 9, 14)

        def _bloque_f(fila, lunes_b):
            ws.cell(fila, 2).value = "Consultor"
            ws.cell(fila, 12).value = "Consultor"
            for i, col in enumerate(cols_dia):
                ws.cell(fila, col).value = _fecha_excel(lunes_b + timedelta(days=i))
            for fila_off in (1, 2):                  # 2 filas → el FESTIVO se combina
                r = fila + fila_off
                ws.cell(r, 2).value = "Mateo Florez"
                ws.cell(r, 12).value = "Mateo Florez"
                for col in cols_dia:                 # cliente en TODAS las columnas-día
                    ws.cell(r, col).value = "HOMI Oracle"

        _bloque_f(1, date(2026, 3, 16))
        _bloque_f(13, date(2026, 3, 23))
        wb.save(ruta)

        out = Path(tmp3) / "out_fest.xlsx"
        actualizar_control_pmp(str(ruta), date(2026, 3, 30), [], cfg, str(out))
        wsf = load_workbook(out, data_only=False)["2026"]
        # Bloque nuevo: header en fila 25, datos en 26-27.
        assert wsf.cell(26, 4).value == "HOMI Oracle"   # lun 30/03 no festivo
        assert wsf.cell(26, 6).value == "HOMI Oracle"   # mar 31/03 no festivo
        assert wsf.cell(26, 8).value == "HOMI Oracle"   # mié 01/04 no festivo
        assert wsf.cell(26, 9).value == "FESTIVO", wsf.cell(26, 9).value   # jue 02/04
        assert wsf.cell(26, 14).value == "FESTIVO", wsf.cell(26, 14).value  # vie 03/04
        # El FESTIVO del jueves se combinó verticalmente sobre las 2 filas de datos.
        assert any(m.min_col == 9 == m.max_col and m.min_row == 26 and m.max_row == 27
                   for m in wsf.merged_cells.ranges), "jueves FESTIVO sin merge vertical"

    print("self-test ok")


# ══════════════════════════════════════════════════════════════════════════════
#  6 · MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    try:
        cfg = cargar_cfg()
        if "--self-test" in sys.argv:
            _selftest()
            sys.exit(0)
        # Herramienta interactiva: necesita una terminal real. Al abrir el
        # ejecutable por doble clic, Windows (PMP.exe) y macOS ('Lanzar PMP.command')
        # la proveen. Si por algún motivo no la hay, se avisa en vez de fallar.
        if not (sys.stdin.isatty() and sys.stdout.isatty()):
            print("Esta herramienta es interactiva: ábrela con doble clic\n"
                  "  · Windows: PMP.exe\n"
                  "  · macOS:   'Lanzar PMP.command'")
            sys.exit(1)
        banner()
        menu(cfg)
    except (KeyboardInterrupt, EOFError):
        console.print("\n  [dim]Interrumpido. Hasta luego 👋[/]\n")
        sys.exit(0)
